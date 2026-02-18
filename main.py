import asyncio
import logging
import re
import sys
import os
from contextlib import asynccontextmanager
from datetime import datetime

from dotenv import load_dotenv
from fastapi import FastAPI, Request, Query, HTTPException
from fastapi.responses import PlainTextResponse
import httpx
from openpyxl import load_workbook

load_dotenv(".env.local")

# ──────────────────────────────────────────────
# Logging – structured output visible in Render
# ──────────────────────────────────────────────
logging.basicConfig(
    stream=sys.stdout,
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("whatsapp-bot")

# ──────────────────────────────────────────────
# Constants – loaded from .env.local
# ──────────────────────────────────────────────
VERIFY_TOKEN = os.getenv("VERIFY_TOKEN", "")
ACCESS_TOKEN = os.getenv("ACCESS_TOKEN", "")
PHONE_NUMBER_ID = os.getenv("PHONE_NUMBER_ID", "")

GRAPH_API_URL = (
    f"https://graph.facebook.com/v21.0/{PHONE_NUMBER_ID}/messages"
)

BOOKING_FILE = os.path.join(os.path.dirname(__file__), "Kapila booking.xlsx")


# ──────────────────────────────────────────────
# Excel-based room availability
# ──────────────────────────────────────────────
ROOM_COLUMNS = ["Room 1", "Room 2", "Room 3", "Room 4", "Room 5"]

DATE_PATTERNS = [
    r"\b(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})\b",     # dd/mm/yyyy or dd-mm-yyyy
    r"\b(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2})\b",       # dd/mm/yy
    r"\b(\d{1,2})\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*"
    r"\s*(\d{4})?\b",                                     # 14 feb or 14 feb 2026
    r"\b(\d{1,2})\s*(st|nd|rd|th)\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*"
    r"\s*(\d{4})?\b",                                     # 14th feb 2026
]

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def parse_date(text: str) -> datetime | None:
    """Try to extract a date from free-form user text."""
    t = text.lower().strip()
    current_year = datetime.now().year

    # dd/mm/yyyy or dd-mm-yyyy or dd.mm.yyyy
    m = re.search(DATE_PATTERNS[0], t)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    # dd/mm/yy
    m = re.search(DATE_PATTERNS[1], t)
    if m:
        try:
            yr = int(m.group(3))
            yr = yr + 2000 if yr < 100 else yr
            return datetime(yr, int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    # 14th feb 2026 (with ordinal suffix)
    m = re.search(DATE_PATTERNS[3], t)
    if m:
        try:
            day = int(m.group(1))
            mon = MONTH_MAP.get(m.group(3)[:3], 0)
            yr = int(m.group(4)) if m.group(4) else current_year
            return datetime(yr, mon, day)
        except (ValueError, KeyError):
            pass

    # 14 feb or 14 feb 2026 (without ordinal)
    m = re.search(DATE_PATTERNS[2], t)
    if m:
        try:
            day = int(m.group(1))
            mon = MONTH_MAP.get(m.group(2)[:3], 0)
            yr = int(m.group(3)) if m.group(3) else current_year
            return datetime(yr, mon, day)
        except (ValueError, KeyError):
            pass

    return None


def check_availability(target_date: datetime) -> str:
    """
    Look up target_date in the Excel sheet and return an availability message.
    """
    date_str = target_date.strftime("%d-%m-%Y")
    display = target_date.strftime("%d %b %Y")

    try:
        wb = load_workbook(BOOKING_FILE, read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        room_indices = [headers.index(r) for r in ROOM_COLUMNS if r in headers]

        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_date = row[0]
            if cell_date is None:
                continue

            if isinstance(cell_date, datetime):
                row_date_str = cell_date.strftime("%d-%m-%Y")
            else:
                row_date_str = str(cell_date).strip()

            if row_date_str == date_str:
                tariff_idx = headers.index("Tariff") if "Tariff" in headers else None
                tariff = row[tariff_idx] if tariff_idx is not None else None

                empty_rooms = sum(1 for i in room_indices if row[i] is None)
                booked_rooms = len(room_indices) - empty_rooms

                logger.info(
                    "availability  | date=%s | booked=%s | empty=%s",
                    date_str, booked_rooms, empty_rooms,
                )

                if empty_rooms > 0:
                    tariff_line = (
                        f"💰 Tariff: *₹{tariff:,.0f}* per room/night\n"
                        if tariff else ""
                    )
                    return (
                        f"✅ *Rooms available on {display}!*\n\n"
                        f"🛏 Available: *{empty_rooms}* of 5 rooms\n"
                        f"📌 Booked: {booked_rooms} of 5\n"
                        f"{tariff_line}\n"
                        "To book, please share:\n"
                        "1️⃣ Number of guests\n"
                        "2️⃣ Number of rooms\n"
                        "3️⃣ Traveling with pets?\n\n"
                        "Or type *book* for full booking details."
                    )
                else:
                    return (
                        f"❌ *Sorry, fully booked on {display}.*\n\n"
                        "All 5 rooms are occupied on this date.\n\n"
                        "💡 Try a nearby date or contact reception:\n"
                        "📞 *+91-XXXXX-XXXXX*"
                    )

        wb.close()
        return (
            f"📅 *{display}*\n\n"
            "We don't have this date in our booking sheet yet.\n"
            "Please contact reception for availability:\n"
            "📞 *+91-XXXXX-XXXXX*"
        )

    except FileNotFoundError:
        logger.error("availability  | file not found: %s", BOOKING_FILE)
        return (
            "⚠️ Booking data is currently unavailable.\n"
            "Please contact reception directly:\n"
            "📞 *+91-XXXXX-XXXXX*"
        )
    except Exception as exc:
        logger.exception("availability  | error reading Excel: %s", exc)
        return (
            "⚠️ Something went wrong while checking availability.\n"
            "Please contact reception:\n"
            "📞 *+91-XXXXX-XXXXX*"
        )


RENDER_URL = os.getenv("RENDER_EXTERNAL_URL", "")
PING_INTERVAL = 14 * 60  # 14 minutes


# ──────────────────────────────────────────────
# Self-ping to keep Render awake
# ──────────────────────────────────────────────
async def keep_alive() -> None:
    """Ping our own /ping endpoint every 14 minutes so Render doesn't sleep."""
    if not RENDER_URL:
        logger.warning("keep_alive    | RENDER_EXTERNAL_URL not set – self-ping disabled")
        return

    url = f"{RENDER_URL}/ping"
    logger.info("keep_alive    | will ping %s every %ss", url, PING_INTERVAL)

    while True:
        await asyncio.sleep(PING_INTERVAL)
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.get(url, timeout=10)
                logger.info("keep_alive    | pinged %s – status %s", url, resp.status_code)
        except Exception as exc:
            logger.error("keep_alive    | ping failed: %s", exc)


@asynccontextmanager
async def lifespan(application: FastAPI):
    task = asyncio.create_task(keep_alive())
    yield
    task.cancel()


app = FastAPI(title="WhatsApp Enquiry Bot", lifespan=lifespan)


@app.get("/ping")
async def ping():
    """Health-check endpoint used by the self-ping task and uptime monitors."""
    return {"status": "alive"}


# ──────────────────────────────────────────────
# Rule-based reply generator
# ──────────────────────────────────────────────
def generate_reply(text: str) -> str:
    """Return a reply based on keyword matching against real Kapila resort info."""
    t = text.lower()

    # ── Greeting ──
    if any(w in t for w in ("hi", "hello", "hey", "hii")):
        return (
            "Welcome to *Kapila River Front*! 🌿🏨\n"
            "A Luxury Farm Villa on the Riverside\n\n"
            "Here's what I can help you with:\n"
            "🛏 *room* – Room details\n"
            "💰 *price* – 2026 Rate card\n"
            "📅 *book* – Booking enquiry\n"
            "🐾 *pet* – Pet policy & charges\n"
            "🎯 *activities* – Sports & games\n"
            "🌿 *amenities* – Facilities\n"
            "🍽 *food* – Dining & meals\n"
            "📍 *location* – How to reach us\n"
            "❌ *cancel* – Cancellation policy\n"
            "🏦 *payment* – Bank & payment info\n"
            "📋 *menu* – Full keyword list\n\n"
            "Or just type your question! 😊"
        )

    # ── Pricing / Rate card ──
    if any(w in t for w in ("price", "cost", "rate", "tariff", "charge", "fee")):
        return (
            "💰 *Kapila River Front – 2026 Rate Card*\n\n"
            "All rates are *per room, per night* for "
            "*double occupancy*, *inclusive of all meals* "
            "(welcome drinks, lunch, high tea, dinner & breakfast).\n\n"
            "📌 *Regular (Non-Seasonal):*\n"
            "• Weekdays: *₹10,000*\n"
            "• Weekends / Holidays: *₹12,000*\n\n"
            "📌 *March – May:*\n"
            "• Weekdays: *₹12,000*\n"
            "• Weekends: *₹13,000*\n\n"
            "📌 *Dasara (10-day festival):*\n"
            "• All days: *₹13,000*\n\n"
            "📌 *December 1–15:*\n"
            "• All days: *₹13,000*\n\n"
            "📌 *December 15–29:*\n"
            "• All days: *₹14,000*\n\n"
            "📌 *January 2 – First weekend:*\n"
            "• All days: *₹14,000*\n\n"
            "📌 *January 5–15:*\n"
            "• All days: *₹12,000*\n\n"
            "📌 *Valentine's Day (14 Feb):*\n"
            "• *₹14,000* per night\n\n"
            "Type *newyear* for the NYE special package!\n"
            "Type *pet* for pet charges or *book* to enquire."
        )

    # ── New Year special package ──
    if any(w in t for w in ("new year", "newyear", "nye", "31 dec", "31st dec",
                             "30 dec", "30th dec", "1 jan", "1st jan", "gala")):
        return (
            "🎆 *New Year Special Package 2026*\n\n"
            "✨ *Mandatory Full Property Booking*\n"
            "📅 2 Nights / 3 Days\n\n"
            "*Option 1:*\n"
            "• Check-in: 30th December\n"
            "• Check-out: 1st January\n\n"
            "*Option 2:*\n"
            "• Check-in: 31st December\n"
            "• Check-out: 2nd January\n\n"
            "💰 *Total Package: ₹2,25,000*\n\n"
            "✅ *Includes:*\n"
            "• All 5 rooms (10 pax)\n"
            "• All meals included\n"
            "• Firecrackers\n"
            "• *New Year Gala Dinner with Barbecue* 🥂\n\n"
            "⚠️ Non-divisible – must be booked as "
            "a full property buyout.\n\n"
            "Type *book* to enquire or *price* for the full rate card."
        )

    # ── Room details ──
    if any(w in t for w in ("room", "bed", "stay", "accommodation", "villa")):
        return (
            "🏨 *Kapila River Front – Room Details*\n\n"
            "We have *5 identical Heritage Rooms*.\n\n"
            "✨ *Room highlights:*\n"
            "• Spacious high-ceiling interior with warm wooden furnishings\n"
            "• Handcrafted wooden bed with elegant ambient lighting\n"
            "• Patterned flooring & tasteful wall art\n"
            "• Private balcony sit-out with comfortable seating\n"
            "• Large glass doors – seamless indoor-outdoor flow\n"
            "• Attached modern washroom with modern fittings\n\n"
            "🔌 *In-room facilities:*\n"
            "• Air Conditioning (A/C)\n"
            "• Television (TV)\n"
            "• Hot water kettle\n\n"
            "👥 *Occupancy:*\n"
            "• Min 2 / Max 3 guests per room\n"
            "• Total: 5 rooms → up to 15 guests (with extra beds)\n\n"
            "ℹ️ One room type only. No river-facing view.\n\n"
            "Type *price* for rates or *book* to enquire!"
        )

    # ── Booking enquiry ──
    if any(w in t for w in ("book", "reserve", "available", "availability",
                             "checkin", "check-in", "checkout", "check-out")):
        return (
            "📅 *Booking Enquiry*\n\n"
            "We'd love to host you at Kapila River Front! 🌿\n\n"
            "🕐 *Check-in:* 1:00 PM\n"
            "🕚 *Check-out:* 11:00 AM\n\n"
            "🔍 *Check availability instantly!*\n"
            "Just send a date like:\n"
            "• *20 mar 2026*\n"
            "• *15/04/2026*\n"
            "• *25th may 2026*\n\n"
            "Or share your booking details:\n"
            "1️⃣ Check-in date\n"
            "2️⃣ Check-out date\n"
            "3️⃣ Number of guests\n"
            "4️⃣ Number of rooms needed\n"
            "5️⃣ Traveling with pets? (Yes/No)\n\n"
            "📞 Or call reception: *+91-XXXXX-XXXXX*\n\n"
            "✅ Booking is confirmed only after *100% payment*.\n"
            "Type *cancel* for cancellation policy.\n"
            "Type *payment* for bank details."
        )

    # ── Pet policy ──
    if any(w in t for w in ("pet", "dog", "cat", "puppy", "animal")):
        return (
            "🐾 *Pet Policy – Kapila River Front*\n\n"
            "Yes! We *welcome pets* and allow them *inside rooms*. 🐶\n\n"
            "📌 *Pet Limits:*\n"
            "• Max *2 pets per room*\n"
            "• Max *6 pets across all 5 rooms* "
            "(if one or more are small breeds like Shih Tzu)\n\n"
            "💰 *Pet Charges:*\n"
            "• *₹2,000 per pet* – includes boiled vegetables & cooked rice\n"
            "• *₹500 extra per pet* – for chicken add-on 🍗\n\n"
            "⚠️ *Guidelines:*\n"
            "• Inform the reservation team *in advance*\n"
            "• Pets must be *leashed if not fully trained*\n"
            "• Owners are *fully responsible* for pet behavior\n"
            "• The property is *open to the riverfront* with no barricading "
            "– please *supervise pets closely* near the river\n"
            "• Any damage or extra cleaning will be *charged to the guest*\n\n"
            "Type *book* to make a reservation or *price* for rates."
        )

    # ── Cancellation policy ──
    if any(w in t for w in ("cancel", "cancellation", "refund", "policy")):
        return (
            "❌ *Cancellation Policy*\n\n"
            "✅ Booking is confirmed only after *100% payment*.\n\n"
            "📌 *Refund rules:*\n"
            "• *15+ days* before check-in → *Full refund* (free cancellation)\n"
            "• *14–15 days* before → *25% deducted*\n"
            "• *10 days* before → *50% deducted*\n"
            "• *Less than 7 days* → *No refund*\n\n"
            "For any changes to your booking, please contact reception:\n"
            "📞 *+91-XXXXX-XXXXX*"
        )

    # ── Payment / bank details ──
    if any(w in t for w in ("payment", "pay", "bank", "account", "upi",
                             "transfer", "neft", "imps", "ifsc")):
        return (
            "🏦 *Payment Details*\n\n"
            "Please transfer to the following account:\n\n"
            "🏛 *Bank:* CANARA BANK\n"
            "👤 *Account Name:* KAPILA RIVER FRONT\n"
            "🔢 *Account Number:* 120032425830\n"
            "🏷 *IFSC Code:* CNRB0002655\n"
            "📍 *Branch:* Ramakrishna Nagar, Mysore\n\n"
            "✅ Booking is confirmed only after *100% payment*.\n\n"
            "After payment, please share the screenshot here "
            "or send it to our reception.\n"
            "📞 *+91-XXXXX-XXXXX*"
        )

    # ── Outdoor sports & activities ──
    if any(w in t for w in ("activit", "sport", "outdoor", "play", "game", "cricket",
                             "badminton", "football", "basketball", "archery", "cycling",
                             "indoor", "table tennis", "foosball", "carrom", "chess")):
        return (
            "🎯 *Kapila River Front – Activities*\n\n"
            "🏏 *Outdoor Sports:*\n"
            "• Netted Cricket\n"
            "• Badminton\n"
            "• Football\n"
            "• Basketball\n"
            "• Archery\n"
            "• Cycling\n\n"
            "🎲 *Indoor Games:*\n"
            "• Table Tennis\n"
            "• Foosball\n"
            "• Carrom\n"
            "• Chess\n"
            "• Puzzle Games\n\n"
            "🏊 *Recreation:*\n"
            "• Swimming Pool\n"
            "• Music System\n\n"
            "✅ All activities are *included* with your stay!\n\n"
            "Type *pool* for swimming pool details."
        )

    # ── Amenities ──
    if any(w in t for w in ("amenit", "facilit", "include", "provide", "offer")):
        return (
            "🌿 *Kapila River Front – Amenities*\n\n"
            "🏠 *In-Room:*\n"
            "• Air Conditioning\n"
            "• Television\n"
            "• Hot water kettle\n"
            "• Attached modern washroom\n"
            "• Private balcony sit-out\n\n"
            "🏟 *On-Site:*\n"
            "• Swimming Pool\n"
            "• Netted Cricket, Badminton, Football, Basketball\n"
            "• Archery & Cycling\n"
            "• Table Tennis, Foosball, Carrom, Chess\n"
            "• Music System\n\n"
            "🍽 *Included:*\n"
            "• All meals (welcome drinks, lunch, high tea, dinner & breakfast)\n"
            "• Peaceful riverside setting\n"
            "• Heritage-style architecture\n\n"
            "Type *activities* for the full list or *price* for rates."
        )

    # ── Swimming pool ──
    if any(w in t for w in ("pool", "swim", "swimming")):
        return (
            "🏊 *Swimming Pool*\n\n"
            "Yes! We have a swimming pool on-site. 💦\n\n"
            "• Accessible to all in-house guests\n"
            "• Included with your stay – no extra charge\n"
            "• Perfect for a refreshing dip after outdoor sports!\n\n"
            "Type *activities* to see all the fun things to do."
        )

    # ── Location / directions ──
    if any(w in t for w in ("location", "address", "direction", "where",
                             "reach", "map", "route", "mysore", "mysuru")):
        return (
            "📍 *How to Reach Kapila River Front*\n\n"
            "Kapila River Front is a luxury farm villa "
            "on the riverside near Mysore.\n\n"
            "📌 For exact location & Google Maps pin, "
            "please contact our reception:\n"
            "📞 *+91-XXXXX-XXXXX*\n\n"
            "We'll share the directions right away! 🗺"
        )

    # ── Food / dining ──
    if any(w in t for w in ("food", "meal", "breakfast", "lunch", "dinner",
                             "dining", "eat", "restaurant", "tea", "drink")):
        return (
            "🍽 *Dining at Kapila River Front*\n\n"
            "All meals are *included* with your stay:\n\n"
            "☕ Welcome drinks on arrival\n"
            "🍛 Lunch\n"
            "🍵 High tea / evening snacks\n"
            "🍽 Dinner\n"
            "🥞 Breakfast (next morning)\n\n"
            "For special dietary needs or meal preferences, "
            "please inform reception in advance:\n"
            "📞 *+91-XXXXX-XXXXX*"
        )

    # ── Valentine's Day ──
    if any(w in t for w in ("valentine", "14 feb", "14th feb")):
        return (
            "💝 *Valentine's Day Special – 14th February*\n\n"
            "🛏 *₹14,000 per night*\n"
            "• Double occupancy\n"
            "• All meals included\n\n"
            "A perfect romantic riverside getaway! 🌹\n\n"
            "Type *book* to reserve or *price* for the full rate card."
        )

    # ── Dasara ──
    if any(w in t for w in ("dasara", "dussehra", "october fest")):
        return (
            "🎆 *Dasara Festival Rates*\n\n"
            "During the *10-day Dasara festival*:\n"
            "• *₹13,000 per night* (all days)\n"
            "• All meals included\n\n"
            "Type *book* to reserve or *price* for the full rate card."
        )

    # ── Thank you / bye ──
    if any(w in t for w in ("thank", "thanks", "bye", "goodbye", "see you")):
        return (
            "Thank you for choosing *Kapila River Front*! 🙏🌿\n\n"
            "We look forward to hosting you.\n"
            "Feel free to message anytime!\n\n"
            "Have a wonderful day! 😊"
        )

    # ── Menu / help ──
    if any(w in t for w in ("menu", "help", "option", "what can")):
        return (
            "📋 *Here's everything I can help with:*\n\n"
            "🛏 *room* – Room details & features\n"
            "💰 *price* – 2026 Rate card\n"
            "🎆 *newyear* – NYE special package\n"
            "💝 *valentine* – Valentine's Day offer\n"
            "📅 *book* – Booking enquiry\n"
            "🐾 *pet* – Pet policy & charges\n"
            "🎯 *activities* – Sports & games\n"
            "🌿 *amenities* – Facilities overview\n"
            "🏊 *pool* – Swimming pool info\n"
            "🍽 *food* – Dining & meals\n"
            "📍 *location* – How to reach us\n"
            "❌ *cancel* – Cancellation policy\n"
            "🏦 *payment* – Bank details\n"
            "👨‍💼 *reception* – Talk to a person\n\n"
            "Just type any keyword! 😊"
        )

    # ── Default fallback ──
    return (
        "Thank you for reaching out to "
        "*Kapila River Front*! 🌿\n\n"
        "I can help you with:\n"
        "🛏 *room* – Room info\n"
        "💰 *price* – 2026 Rates\n"
        "📅 *book* – Booking enquiry\n"
        "🐾 *pet* – Pet policy\n"
        "🎯 *activities* – Things to do\n"
        "❌ *cancel* – Cancellation policy\n"
        "📋 *menu* – See all options\n\n"
        "Or type your question and our team "
        "will get back to you! 🙏"
    )


# ──────────────────────────────────────────────
# Send a WhatsApp message via the Graph API
# ──────────────────────────────────────────────
async def send_message(to: str, message: str) -> None:
    """Send a text message through the Meta WhatsApp Cloud API."""
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": to,
        "type": "text",
        "text": {"body": message},
    }

    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                GRAPH_API_URL, headers=headers, json=payload
            )
            logger.info("send_message  | to=%s | status=%s", to, response.status_code)
            logger.info("send_message  | response=%s", response.text)
        except httpx.RequestError as exc:
            logger.error("send_message  | request failed: %s", exc)


# ──────────────────────────────────────────────
# Send interactive button menus via the Graph API
# ──────────────────────────────────────────────
async def _send_interactive(to: str, body_text: str, buttons: list[dict]) -> None:
    """Low-level helper to send any interactive button message."""
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": to,
        "type": "interactive",
        "interactive": {
            "type": "button",
            "body": {"text": body_text},
            "action": {"buttons": buttons},
        },
    }

    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                GRAPH_API_URL, headers=headers, json=payload
            )
            logger.info("send_buttons  | to=%s | status=%s", to, response.status_code)
            logger.info("send_buttons  | response=%s", response.text)
        except httpx.RequestError as exc:
            logger.error("send_buttons  | request failed: %s", exc)


async def send_button_message(to: str) -> None:
    """Main welcome menu – shown on greeting."""
    await _send_interactive(
        to,
        "Welcome to *Kapila River Front*! 🌿🏨\n"
        "A Luxury Farm Villa on the Riverside\n\n"
        "How may I assist you today?",
        [
            {"type": "reply", "reply": {"id": "room", "title": "Room Details 🛏"}},
            {"type": "reply", "reply": {"id": "price", "title": "2026 Rate Card 💰"}},
            {"type": "reply", "reply": {"id": "more", "title": "More Options 📋"}},
        ],
    )


async def send_more_options(to: str) -> None:
    """Second menu – activities, pets, policies."""
    await _send_interactive(
        to,
        "More about *Kapila River Front* 🌿",
        [
            {"type": "reply", "reply": {"id": "activities", "title": "Activities 🎯"}},
            {"type": "reply", "reply": {"id": "pet", "title": "Pet Policy 🐾"}},
            {"type": "reply", "reply": {"id": "policies", "title": "Policies 📄"}},
        ],
    )


async def send_policies_menu(to: str) -> None:
    """Third menu – cancellation, payment, reception."""
    await _send_interactive(
        to,
        "📄 *Booking & Policies*",
        [
            {"type": "reply", "reply": {"id": "cancel", "title": "Cancellation ❌"}},
            {"type": "reply", "reply": {"id": "payment", "title": "Payment Info 🏦"}},
            {"type": "reply", "reply": {"id": "reception", "title": "Reception 👨‍💼"}},
        ],
    )


# ──────────────────────────────────────────────
# Handle interactive button clicks
# ──────────────────────────────────────────────
async def handle_button_click(sender: str, button_id: str) -> None:
    """Route logic based on the button ID the user tapped."""
    logger.info("button_click  | from=%s | button_id=%s", sender, button_id)

    if button_id == "price":
        await send_message(sender, generate_reply("price"))
        await send_button_message(sender)

    elif button_id == "room":
        await send_message(sender, generate_reply("room"))
        await send_button_message(sender)

    elif button_id == "activities":
        await send_message(sender, generate_reply("activities"))
        await send_more_options(sender)

    elif button_id == "pet":
        await send_message(sender, generate_reply("pet"))
        await send_more_options(sender)

    elif button_id == "cancel":
        await send_message(sender, generate_reply("cancel"))
        await send_policies_menu(sender)

    elif button_id == "payment":
        await send_message(sender, generate_reply("payment"))
        await send_policies_menu(sender)

    elif button_id == "more":
        await send_more_options(sender)

    elif button_id == "policies":
        await send_policies_menu(sender)

    elif button_id == "reception":
        await send_message(
            sender,
            "👨‍💼 *Connecting you to our reception!*\n\n"
            "📞 Call us: *+91-XXXXX-XXXXX*\n"
            "💬 WhatsApp: *+91-XXXXX-XXXXX*\n\n"
            "Our team (Prajwal – Reservation Team) "
            "will assist you right away! 🙏"
        )

    else:
        await send_message(sender, generate_reply(""))
        await send_button_message(sender)


# ──────────────────────────────────────────────
# Webhook verification (GET)
# ──────────────────────────────────────────────
@app.get("/webhook")
async def verify_webhook(
    hub_mode: str = Query(None, alias="hub.mode"),
    hub_verify_token: str = Query(None, alias="hub.verify_token"),
    hub_challenge: str = Query(None, alias="hub.challenge"),
):
    """
    Meta sends a GET request with hub.mode, hub.verify_token, and
    hub.challenge to verify the webhook URL.
    """
    if hub_mode == "subscribe" and hub_verify_token == VERIFY_TOKEN:
        logger.info("verify        | webhook verified successfully")
        return PlainTextResponse(content=hub_challenge)

    logger.warning("verify        | verification failed – token mismatch")
    raise HTTPException(status_code=403, detail="Verification failed")


# ──────────────────────────────────────────────
# Webhook receiver (POST)
# ──────────────────────────────────────────────
@app.post("/webhook")
async def receive_webhook(request: Request):
    """
    Receives incoming messages from WhatsApp, generates a reply,
    and sends it back to the sender.
    """
    body = await request.json()
    logger.info("webhook       | incoming payload: %s", body)

    try:
        entry = body.get("entry", [])
        for e in entry:
            changes = e.get("changes", [])
            for change in changes:
                value = change.get("value", {})
                messages = value.get("messages", [])

                for msg in messages:
                    sender = msg.get("from")
                    msg_type = msg.get("type")

                    if msg_type == "interactive":
                        button_id = (
                            msg.get("interactive", {})
                            .get("button_reply", {})
                            .get("id", "")
                        )
                        await handle_button_click(sender, button_id)

                    elif msg_type == "text":
                        text = msg.get("text", {}).get("body", "")
                        logger.info("webhook       | from=%s | text=%s", sender, text)

                        greetings = ("hi", "hello", "hey", "hii", "helo",
                                     "good morning", "good afternoon",
                                     "good evening")

                        parsed = parse_date(text)

                        if text.lower().strip() in greetings:
                            await send_button_message(sender)
                        elif parsed is not None:
                            logger.info("webhook       | date detected: %s", parsed.strftime("%d-%m-%Y"))
                            reply = check_availability(parsed)
                            await send_message(sender, reply)
                        elif any(w in text.lower() for w in ("menu", "help", "option")):
                            await send_message(sender, generate_reply("menu"))
                            await send_button_message(sender)
                        else:
                            reply = generate_reply(text)
                            await send_message(sender, reply)

                    else:
                        logger.info("webhook       | from=%s | unhandled type=%s", sender, msg_type)
                        await send_message(sender, generate_reply(""))
                        await send_button_message(sender)

                    logger.info("webhook       | replied to=%s", sender)

    except Exception as exc:
        logger.exception("webhook       | error processing message: %s", exc)

    return {"status": "ok"}
